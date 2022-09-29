from os import getenv
import pymssql
import os
import pandas as pd
import codecs
import subprocess

from openpyxl import load_workbook

wb=load_workbook('worksheet4.xlsx')
sheet=wb.worksheets[0]

row=sheet.max_row-1
print(row)

xl = pd. ExcelFile("worksheet4.xlsx")

df = xl.parse("DB Deployment")



diction={}
keys=range(row)
for i in keys:
         diction[i]=(df[ 'Object SQL File Name'][i],df['Change set #'][i],df['Destination Database IP Address'][i],df['Destination Database Name'][i],df['Port#'][i],df['Object Type (SP/FUNC/TABLE/Data Script)'][i],df['Is this New or Old Object'][i])
         
print(diction)

listSP=("SP","sp","Sp","StoredProcedures","Stored Procedures","StoredProcedure","Stored Procedure","Stored procedure","stored procedures","stored procedure","STORED PROCEDURES","STOREDPROCEDURES","STORED PROCEDURE","STOREPROCEDURE")
neworold=("old","Old","OLD")
for i in keys: 
         if((df['Object Type (SP/FUNC/TABLE/Data Script)'][i] in listSP) and (df['Is this New or Old Object'][i] in neworold)):
                                    filename=diction[i][0]
                                    print(filename)
                                    command1="sp_helptext "+filename[:-4]
                                    
                                    filename1="saveit.sql"
                                    f1=open(filename1,"w")
                                    f1.write(command1)
                                    f1.close()
                                    folder=diction[i][3]              #AxisMF
                                    path=".\\"+folder               #.\AxisMF

                                    if os.path.isdir(path):
                                                           path1=".\\"+folder+"\\"+filename   #.\AxisMF\ABC.sql
             
                                    else:
                                                           os.mkdir(path)
                                                           path1=".\\"+folder+"\\"+filename   #.\AxisMF\ABC.sql

                                    servername=diction[i][2]+","+str(diction[i][4])

                                    command2="sqlcmd -S "+servername+" -d "+diction[i][3]+" -e -i "+filename1+" -o "+path1  #sqlcmd -S D1WB3WV-4155 -d DBJenkins -U so -P Unlock@123 -i saveit.sql -o .\DBJenkins\SelectAllCustomers.sql
                                    print(command2)
                                    f2=open("batchfile.bat","w")
                                    f2.write(command2)
                                    f2.close()
                                    subprocess.call([r'batchfile.bat'])
                                    """f3=open(path1,"w")

                                    line1=f3.readline()[0:30]
                                    line2=f3.readline()[0:]
                                    t="Text"
                                    if((line1==command1) & line2==t):
                                                           f4=open("Result.txt","w")
                                                           penaldo=filename+folder+" success"
                                                           f4.write(penaldo)
                                                           f4.write("\n")

                                    else:
                                                           f4=open("Result.txt","w")
                                                           penaldo=filename+folder+" failure"
                                                           f4.write(penaldo)
                                                           f4.write("\n")                       


                                     """                                           
                                    os.remove("batchfile.bat")
                                    os.remove("saveit.sql")
                                    #f4.close()
                                    #f3.close()
         else:
                                    print("New SP or table")
                                    continue




   