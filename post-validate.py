from os import getenv
import pymssql
import os
import pandas as pd
import codecs
import subprocess

from openpyxl import load_workbook

wb=load_workbook('worksheet4.xlsx')
sheet=wb.worksheets[0]

row=sheet.max_row-1
print(row)

xl = pd. ExcelFile("worksheet4.xlsx")

df = xl.parse("DB Deployment")



diction={}
keys=range(row)
for i in keys:
         diction[i]=(df[ 'Object SQL File Name'][i],df['Change set #'][i],df['Destination Database IP Address'][i],df['Destination Database Name'][i],df['Port#'][i],df['Object Type (SP/FUNC/TABLE/Data Script)'][i],df['Is this New or Old Object'][i])
         
print(diction)

listSP=("SP","sp","Sp","StoredProcedures","Stored Procedures","StoredProcedure","Stored Procedure","Stored procedure","stored procedures","stored procedure","STORED PROCEDURES","STOREDPROCEDURES","STORED PROCEDURE","STOREPROCEDURE")
neworold=("old","Old","OLD")



for i in keys:
         

 if(diction[i][6] in neworold):  
         filename5=diction[i][0]
         command11="sp_helptext "+filename5[:-4]
         l=len(command11)
         folder1=diction[i][3] 
         path3=folder1+"\\"+filename5
         
         f3=open(path3,"r")

         line1=f3.readline()[0:l]
         line2=f3.readline()[0:4]
         line3=f3.readline()[l:100]
         line4=line1+line3
         
         print(line1)
         
         t="Text"
         if((line1==command11) & (line2==t)):
                                f4=open("Result.txt","a")
                                result=str(i)+"."+filename5+" "+folder1+" success"
                                f4.write(result)
                                f4.write("\n \n")
                                f4.close()

         else:
                                f4=open("Result.txt","a")
                                
                               
                                result=str(i)+"."+filename5+" "+folder1+" failure "+line4
                                f4.write(result)
                                f4.write("\n \n")
                                f4.write
                                f4.close()

f4.close()
f3.close()      

os.system("notepad Result.txt")              