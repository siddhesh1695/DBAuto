from os import getenv
import pymssql
import os
import pandas as pd
import codecs
import subprocess
import shutil

from openpyxl import load_workbook

wb=load_workbook('worksheet.xlsx')
sheet=wb.worksheets[1]

row=sheet.max_row-1
print(row)

xl = pd. ExcelFile("worksheet.xlsx")

df = xl.parse("TFS")

diction={}




keys=range(row)
for i in keys:
         diction[i]=(df[ 'Object SQL File Name'][i],df['Change set #'][i],df['Destination Database IP Address'][i],df['Destination Database Name'][i],df['Port#'][i],df['Object Type (SP/FUNC/TABLE/Data Script)'][i],df['TFS Name'][i])
         
print(diction)





listSP=("SP","sp","Sp","StoredProcedures","Stored Procedures","StoredProcedure","Stored Procedure","stored procedures","stored procedure","STORED PROCEDURES","STOREDPROCEDURES","STORED PROCEDURE","STOREPROCEDURE")
listtable=("Tables","tables","TABLES","Table","table","TABLE")


envmnt="UAT"


for i in keys:
         x=diction[i][5]
         if(x in listSP):
                 filename=diction[i][0]
                 object="Stored Procedures"
                 dbname=diction[i][3]
                 tfsname=diction[i][6]
                 path="D:\\DFS_Database\\"+tfsname+"\\"+envmnt+"\\"+object          
                 
                 sourcepath=path+"\\"+filename
                 command1="cd "+path
                 f=open("TFSbatch.bat","w")
                 f.write("D:")
                 f.write("\n")
                 f.write(command1)
                 command2="tf get /version:"+str(diction[i][1])+" "+diction[i][0]
                 f.write("\n")
                 f.write(command2)
                 f.close()
                 subprocess.call([r'TFSbatch.bat'])         
                 curr_dir=os.getcwd()
                 destpath=curr_dir+"\\"+filename
                 shutil.copy(sourcepath, destpath)
                 folder=diction[i][3]              #AxisMF
                 folderpth=".\\"+folder               #.\AxisMF

                 if os.path.isdir(folderpth):
                            penaldo=".\\"+folder+"\\"+filename   #.\AxisMF\ABC.sql
             
                 else:
                            os.mkdir(folderpth)
                            penaldo=".\\"+folder+"\\"+filename   #.\AxisMF\ABC.sql
                 command3="sqlcmd -S "+diction[i][2]+" -d "+diction[i][3]+" -U so -P Unlock@123 -i "+filename+" -o "+penaldo
                 print(command3)
                 f1=open("Execute.bat","w")
                 f1.write(command3)
                 f1.close()
                 subprocess.call([r'Execute.bat'])
                 os.remove(filename)
                 os.remove("Execute.bat")

         elif(x in listtable):
                 filename=diction[i][0]
                 object="Tables"
                 dbname=diction[i][3]
                 tfsname=TFSDB[dbname]
                 path="D:\\DFS_Database\\"+tfsname+"\\"+envmnt+"\\"+object          
                 sourcepath=path+"\\"+filename
                 command1="cd "+path
                 f=open("TFSbatch.bat","w")
                 f.write("D:")
                 f.write("\n")
                 f.write(command1)
                 command2="tf get /version:"+str(diction[i][1])+" "+diction[i][0]
                 f.write("\n")
                 f.write(command2)
                 f.close()
                 subprocess.call([r'TFSbatch.bat'])         
                 curr_dir=os.getcwd()
                 destpath=curr_dir+"\\"+filename
                 shutil.copy(sourcepath, destpath)
                 command3="sqlcmd -S "+diction[i][2]+" -d "+diction[i][3]+" -U so -P Unlock@123 -i "+filename+" -o Output.txt"
                 print(command3)
                 f1=open("Execute.bat","w")
                 f1.write(command3)
                 f1.close()
                 subprocess.call([r'Execute.bat'])
                 os.remove(filename)
                 os.remove("Execute.bat")

         else:
                 print("Nothing to execute")

         
