from os import getenv
import pymssql
import os
import pandas as pd
import codecs
import subprocess

from openpyxl import load_workbook

wb=load_workbook('worksheet4.xlsx')
sheet=wb.worksheets[0]

row=sheet.max_row-1
print(row)

xl = pd. ExcelFile("worksheet4.xlsx")

df = xl.parse("DB Deployment")

"""

diction={}
keys=range(row)
for i in keys:
         diction[i]=(df[ 'Object SQL File Name'][i],df['Change set #'][i],df['TFS Name'][i],df['Destination Database Name'][i],df['Port#'][i],df['Object Type (SP/FUNC/TABLE/Data Script)'][i],df['Is this New or Old Object'][i])
         


listSP=("SP","sp","Sp","StoredProcedures","Stored Procedures","StoredProcedure","Stored Procedure","Stored procedure","stored procedures","stored procedure","STORED PROCEDURES","STOREDPROCEDURES","STORED PROCEDURE","STOREPROCEDURE")

neworold=("old","Old","OLD")

new=("new","New","NEW")

listtable=("Tables","tables","TABLES","Table","table","TABLE")

ports=(14999,15999,16999,17999)


DBName_DFS=("MUSCAT","BARODA","KARVY MFS","SAHARA","BAJAJ","MOTILALOSWAL","KBOLT","UTI","ITI","BNPMF","MIRAE","EDELWEISS","AXA","CANROBECO","TAURUS","LIC","JMMF","SAMCO","SUNDARAMMF","TRUST","QUANTUM","QUANT","RELIANCE","RELIGARE","IDBIMF","IBMF","NJMF","DLFPRAMERICA","PEERLESSMF","AXISMF")

TFSName=("Tailwind","Taurus","Trust","UTI","SAMCO","Sundaram","Sahara","AXISMF","Bajaj","Baroda","BNP","BOIAXA","CRMF","DFS_AllJobs","Edelwiess","HRISINT","HelpDesk","Invesco","Invit","ITI","IndiaBulls","IDBI","JMMF","Karvy_MFS","KBOLT","LICMF","MFDWeb","Mirae","Motilal","Muscat","Nippon","Navi","NJMF","PGIM","Quant","Quantum")

print("\n Validating....\n")

print("\n \n Validating Object Types... ")

for i in keys: 
         
         x=diction[i][5]
         
         if ((x in listSP) or (x in listtable)):
                            print("\n Validation of Object type successful \n")

         else: 
                            print("\n Validation failed for row no"+str(i+2)+"\n")

print("\n \n Validating Old or new...")
for i in keys:

         
         y=diction[i][6]
         if ((y in neworold) or (y in new)):
                            print("\n Validation of Old or new successful\n")

         else:
                            print("\n Validation failed for row no"+str(i+2))





print("\n \n Validating Database names...")
for i in keys:

         
         z=diction[i][3]
         zup=z.upper()
         if (zup in DBName_DFS):
                            print("\n Validation of Database names "+zup+" successful\n")

         else:
                            print("\n Validation of Database names "+zup+" failed of row "+str(i+2)+"\n")
                            


print("\n \n Validating Port nos...\n")
for i in keys:

         
         a=diction[i][4]
         if (a in ports):
                            print("\n Validation of Database Port Nos successful\n")

         else:
                            print("\n Validation failed for row no "+str(i+2))
                            print(" Of port"+str(a))



"""       